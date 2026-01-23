#!/usr/bin/env python3
"""
NADESHIKO 分析.xlsx → CSV変換スクリプト

月別シート（1月〜8月）を個別CSVに変換する。
"""

import openpyxl
import csv
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

# 設定
EXCEL_PATH = Path(__file__).parent.parent / 'data/再生数シート/NADESHIKO 分析.xlsx'
OUTPUT_DIR = Path(__file__).parent.parent / 'data/再生数シート'

# 月別シート設定（シート名: (データ開始行, 年月)）
MONTHLY_SHEETS = {
    '1月': (17, '2026年1月'),
    '12月': (1, '2025年12月'),
    '11月': (1, '2025年11月'),
    '10月': (1, '2025年10月'),
    '9月': (1, '2025年9月'),
    '8月': (1, '2025年8月'),
}


def convert_sheet_to_csv(wb, sheet_name, start_row, output_name):
    """シートをCSVに変換"""
    ws = wb[sheet_name]
    output_path = OUTPUT_DIR / f'{output_name}.csv'

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in range(start_row, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                # 改行を除去（CSVが壊れないように）
                if isinstance(val, str):
                    val = val.replace('\n', ' ').replace('\r', '')
                row_data.append(val if val is not None else '')
            # 空行スキップ（最初のカラムが空の場合）
            if row_data[0]:
                writer.writerow(row_data)

    return output_path


def main():
    print(f'読み込み: {EXCEL_PATH.name}')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    total_rows = 0
    for sheet_name, (start_row, output_name) in MONTHLY_SHEETS.items():
        path = convert_sheet_to_csv(wb, sheet_name, start_row, output_name)
        ws = wb[sheet_name]
        rows = ws.max_row - start_row + 1
        total_rows += rows
        print(f'✓ {sheet_name} → {path.name} ({rows}行)')

    print(f'\n完了: 6ファイル, 合計{total_rows}行')


if __name__ == '__main__':
    main()
