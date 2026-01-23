#!/usr/bin/env python3
"""
過去再生数シートExcel → CSV変換スクリプト

【新】NADESIKO_投稿インサイトDB .xlsx から2025年1月〜7月のデータを
再生数シート/にCSVとして追加する。
"""

import openpyxl
import csv
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

EXCEL_PATH = Path(__file__).parent.parent / 'data/過去再生数シート/【新】NADESIKO_投稿インサイトDB .xlsx'
OUTPUT_DIR = Path(__file__).parent.parent / 'data/再生数シート'

# 対象シート（シート名: (データ開始行, 出力ファイル名)）
MONTHLY_SHEETS = {
    # 2024年
    '【4月】': (7, '2024年4月'),
    '【5月】': (8, '2024年5月'),
    '【6月】': (7, '2024年6月'),
    '【7月】': (12, '2024年7月'),
    '【8月】': (12, '2024年8月'),
    '【9月】': (12, '2024年9月'),
    '【10月】': (12, '2024年10月'),
    '【11月】': (12, '2024年11月'),
    '【12月】': (12, '2024年12月'),
    # 2025年
    '【20257】': (13, '2025年7月'),
    '【20256】': (13, '2025年6月'),
    '【20255】': (13, '2025年5月'),
    '【20254】': (13, '2025年4月'),
    '【20253】': (13, '2025年3月'),
    '【20252】': (13, '2025年2月'),
    '【20251】': (13, '2025年1月'),
}

# 新CSVヘッダー（既存CSVと同じ28列）
NEW_HEADERS = [
    '投稿日', 'アカウント名', 'PR/通常', 'sns', '担当者', 'タイトル', 'URL',
    '種別', '更新日', '動画尺', '再生数', 'いいね', 'コメント', '共有', '保存',
    'いいね率', 'コメント率', '共有率', '保存率', '平均視聴時間', '視聴維持率',
    '1秒継続率', '3秒継続率', '6秒継続率', 'フル視聴率', '新規フォロー', 'フォロー率', 'おすすめ率'
]


def convert_row(ws, row_num):
    """旧Excelの行を新CSV形式に変換"""
    def get_val(col):
        v = ws.cell(row=row_num, column=col).value
        if v is None:
            return ''
        if isinstance(v, str):
            return v.replace('\n', ' ').replace('\r', '')
        return v

    return [
        get_val(1),   # 投稿日
        get_val(2),   # アカウント名
        get_val(5),   # PR/通常 (通常 or PR)
        '',           # sns (空)
        '',           # 担当者 (空)
        get_val(3),   # タイトル
        get_val(4),   # URL
        get_val(6),   # 種別
        '',           # 更新日 (空)
        get_val(8),   # 動画尺
        get_val(9),   # 再生数
        get_val(10),  # いいね
        get_val(11),  # コメント
        get_val(12),  # 共有
        get_val(13),  # 保存
        get_val(14),  # いいね率
        get_val(15),  # コメント率
        get_val(16),  # 共有率
        get_val(17),  # 保存率
        get_val(18),  # 平均視聴時間
        get_val(19),  # 視聴維持率
        get_val(20),  # 1秒継続率
        get_val(21),  # 3秒継続率
        get_val(22),  # 6秒継続率
        get_val(23),  # フル視聴率
        get_val(24),  # 新規フォロー
        get_val(25),  # フォロー率
        get_val(26),  # おすすめ率
    ]


def main():
    print(f'読み込み: {EXCEL_PATH.name}')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    total_rows = 0
    for sheet_name, (start_row, output_name) in MONTHLY_SHEETS.items():
        ws = wb[sheet_name]
        output_path = OUTPUT_DIR / f'{output_name}.csv'

        rows_written = 0
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(NEW_HEADERS)

            for row in range(start_row, ws.max_row + 1):
                first_val = ws.cell(row=row, column=1).value
                if first_val:  # 空行スキップ
                    row_data = convert_row(ws, row)
                    writer.writerow(row_data)
                    rows_written += 1

        total_rows += rows_written
        print(f'✓ {sheet_name} → {output_path.name} ({rows_written}行)')

    print(f'\n完了: {len(MONTHLY_SHEETS)}ファイル, 合計{total_rows}行')


if __name__ == '__main__':
    main()
